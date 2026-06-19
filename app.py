import io
import json
import logging
from flask import Flask, jsonify, request, render_template, Response, send_file
from flask_cors import CORS
from datetime import datetime

# Openpyxl for styled Excel sheet
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import db
import scraper

# Initialize logger
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = Flask(__name__, static_folder='static', template_folder='templates')
CORS(app)

# Ensure DB is initialized
db.init_db()

def get_session_id():
    return request.headers.get('X-Session-Id') or request.args.get('session_id') or 'default'

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/jobs', methods=['GET'])
def get_jobs():
    session_id = get_session_id()
    filters = {
        'status': request.args.get('status'),
        'site': request.args.get('site'),
        'country': request.args.get('country'),
        'search': request.args.get('search')
    }
    
    # parse remote filter
    is_remote_param = request.args.get('is_remote')
    if is_remote_param is not None:
        if is_remote_param.lower() in ['true', '1']:
            filters['is_remote'] = True
        elif is_remote_param.lower() in ['false', '0']:
            filters['is_remote'] = False
            
    jobs = db.get_jobs(filters, session_id=session_id)
    return jsonify(jobs)

@app.route('/api/jobs/<job_id>', methods=['PUT'])
def update_job(job_id):
    session_id = get_session_id()
    data = request.json or {}
    status = data.get('status')
    notes = data.get('notes')
    
    if not status:
        return jsonify({'error': 'Status is required'}), 400
        
    db.update_job_status(job_id, status, notes, session_id=session_id)
    return jsonify({'message': 'Job updated successfully'})

@app.route('/api/stats', methods=['GET'])
def get_stats():
    session_id = get_session_id()
    stats = db.get_stats(session_id=session_id)
    return jsonify(stats)

@app.route('/api/jobs/clear', methods=['POST'])
def clear_jobs():
    session_id = get_session_id()
    try:
        db.clear_all_jobs(session_id=session_id)
        return jsonify({'message': 'All jobs cleared successfully'})
    except Exception as e:
        logger.error(f"Error clearing jobs: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/scrape', methods=['GET', 'POST'])
def run_scrape():
    """
    Triggers scraping and streams output using Server-Sent Events (SSE)
    """
    session_id = get_session_id()
    search_term = request.args.get('search_term', 'Flutter Developer')
    location = request.args.get('location', 'Remote')
    hours_old = int(request.args.get('hours_old', 72))
    results_per_site = int(request.args.get('results_per_site', 25))
    is_remote = request.args.get('is_remote', 'true').lower() in ['true', '1']
    clear_before = request.args.get('clear_before', 'false').lower() in ['true', '1']
    
    sites_param = request.args.get('sites')
    if sites_param:
        sites = [s.strip() for s in sites_param.split(',')]
    else:
        sites = ['linkedin', 'google', 'zip_recruiter', 'indeed']
        
    country_indeed = request.args.get('country_indeed', 'USA')

    def event_stream():
        # Yield initial message
        yield f"data: {json.dumps({'message': 'Connection established. Scraper starting...'})}\n\n"
        
        if clear_before:
            try:
                db.clear_all_jobs(session_id=session_id)
                clear_msg = json.dumps({'message': 'Existing jobs cleared from database.\n'})
                yield f"data: {clear_msg}\n\n"
            except Exception as e:
                err_msg = json.dumps({'message': f'ERROR clearing database: {str(e)}\n'})
                yield f"data: {err_msg}\n\n"
        
        try:
            generator = scraper.scrape_and_save_all(
                search_term=search_term,
                location=location,
                is_remote=is_remote,
                hours_old=hours_old,
                results_per_site=results_per_site,
                sites=sites,
                country_indeed=country_indeed,
                session_id=session_id
            )
            for progress_msg in generator:
                yield f"data: {json.dumps({'message': progress_msg})}\n\n"
                
            yield f"data: {json.dumps({'message': '---SCRAPE_COMPLETE---'})}\n\n"
        except Exception as e:
            logger.error(f"Error in scraping stream: {str(e)}")
            error_msg = f"ERROR: {str(e)}\n"
            yield f"data: {json.dumps({'message': error_msg})}\n\n"
            yield f"data: {json.dumps({'message': '---SCRAPE_FAILED---'})}\n\n"

    return Response(event_stream(), content_type='text/event-stream')

@app.route('/api/export', methods=['GET'])
def export_jobs():
    """
    Generates and downloads a beautifully styled Excel spreadsheet with job details
    and application trackers.
    """
    session_id = get_session_id()
    # Fetch all jobs matching user's current filters
    filters = {
        'status': request.args.get('status'),
        'site': request.args.get('site'),
        'country': request.args.get('country'),
        'search': request.args.get('search')
    }
    is_remote_param = request.args.get('is_remote')
    if is_remote_param is not None:
        if is_remote_param.lower() in ['true', '1']:
            filters['is_remote'] = True
        elif is_remote_param.lower() in ['false', '0']:
            filters['is_remote'] = False
            
    jobs = db.get_jobs(filters, session_id=session_id)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Hunter Jobs"
    
    # Ensure grid lines are visible
    ws.views.sheetView[0].showGridLines = True
    
    # Headers
    headers = [
        "Platform", "Job Title", "Company", "Location", "Country", "Job Type", 
        "Date Posted", "Salary Min", "Salary Max", "Currency", 
        "Remote?", "Application Status", "Date Applied", "Notes", "Job Apply Link"
    ]
    
    # Styling variables
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    
    # Write headers
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    ws.row_dimensions[1].height = 28
    
    # Border styling
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Status coloring palettes
    status_fills = {
        "Applied": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),        # Soft green
        "Interviewing": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),   # Soft blue
        "Offer": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),          # Soft gold
        "Rejected": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),       # Soft red/orange
        "Interested": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),     # Soft grey
        "Not Applied": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    }
    
    status_fonts = {
        "Applied": Font(name=font_family, size=10, bold=True, color="375623"),
        "Interviewing": Font(name=font_family, size=10, bold=True, color="1F4E78"),
        "Offer": Font(name=font_family, size=10, bold=True, color="7F6000"),
        "Rejected": Font(name=font_family, size=10, bold=True, color="C65911"),
        "Interested": Font(name=font_family, size=10, bold=True, color="595959"),
        "Not Applied": Font(name=font_family, size=10, color="000000")
    }

    body_font = Font(name=font_family, size=10)
    link_font = Font(name=font_family, size=10, color="0563C1", underline="single")
    
    for row_idx, job in enumerate(jobs, start=2):
        min_sal = job.get('min_amount')
        max_sal = job.get('max_amount')
        
        row_data = [
            job.get('site', '').upper(),
            job.get('title', ''),
            job.get('company', ''),
            job.get('location', ''),
            job.get('country', ''),
            job.get('job_type', ''),
            job.get('date_posted', ''),
            min_sal if min_sal is not None else "",
            max_sal if max_sal is not None else "",
            job.get('currency', ''),
            "Yes" if job.get('is_remote') else "No",
            job.get('status', 'Not Applied'),
            job.get('date_applied', '') or "",
            job.get('notes', ''),
            "Apply"  # Placeholder value for cell, we set hyperlink below
        ]
        
        ws.append(row_data)
        ws.row_dimensions[row_idx].height = 20
        
        status = job.get('status', 'Not Applied')
        status_fill = status_fills.get(status, status_fills['Not Applied'])
        status_font = status_fonts.get(status, status_fonts['Not Applied'])
        
        # Format columns in the row
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = body_font
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
            
            # Application status column styling
            if col_idx == 12:
                cell.fill = status_fill
                cell.font = status_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply hyperlinking to column 15
            elif col_idx == 15:
                url = job.get('job_url')
                if url:
                    cell.hyperlink = url
                    cell.font = link_font
                    cell.value = "Click to Apply"
                else:
                    cell.value = "N/A"
                cell.alignment = Alignment(horizontal="center", vertical="center")
                
            # Formatting alignments for specific columns
            elif col_idx in [1, 6, 7, 10, 11, 13]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [8, 9]:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if cell.value != "":
                    cell.number_format = '#,##0.00'
                    
    # Auto-adjust column width based on content length
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            val = str(cell.value or '')
            # Don't size column based on hyperlink formula or long URL strings
            if cell.column == 15:
                max_len = max(max_len, len("Click to Apply"))
            elif cell.column == 14: # Notes column
                max_len = max(max_len, min(len(val), 30)) # Caps notes column scaling width
            else:
                max_len = max(max_len, len(val))
                
        ws.column_dimensions[col_letter].width = max(max_len + 3, 11)
        
    # Save file into memory buffer
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"job_hunter_export_{timestamp}.xlsx"
    
    return send_file(
        file_stream,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    # Initialize DB schema
    db.init_db()
    # Run server locally on port 5001
    app.run(debug=True, host='0.0.0.0', port=5001)
